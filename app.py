
import streamlit as st
import pandas as pd
import openpyxl
from io import BytesIO

# --- Constants ---
TEMPLATE_PATH = "sku-template (4).xlsx"
MAPPING_PATH = "Mapping - Automation.xlsx"

# --- Load Mapping Sheet (only from backend) ---
@st.cache_data
def load_mapping():
    df = pd.read_excel(MAPPING_PATH, sheet_name=0)
    return df

def process_file(input_file, mapping_df):
    # Load input and template
    input_df = pd.read_excel(input_file)
    wb = openpyxl.load_workbook(TEMPLATE_PATH)
    ws_values = wb["Values"]
    ws_type = wb["Type"]

    # --- Step 1: Fill "Values" tab ---
    for i, row in enumerate(input_df.itertuples(index=False), start=2):
        for j, value in enumerate(row, start=1):
            ws_values.cell(row=i, column=j, value=value)
    # Paste header
    for j, col_name in enumerate(input_df.columns, start=1):
        ws_values.cell(row=1, column=j, value=col_name)

    # --- Step 2: Fill "Type" tab Row 1 & 2 with header ---
    for col_idx, header in enumerate(input_df.columns, start=2):
        ws_type.cell(row=1, column=col_idx, value=header)
        ws_type.cell(row=2, column=col_idx, value=header)

    # --- Step 3: Map values using Mapping sheet ---
    for col_idx, header in enumerate(input_df.columns, start=2):
        match_row = mapping_df[mapping_df.iloc[:, 1] == header]
        if not match_row.empty:
            row3_val = match_row.iloc[0, 3]  # Column D
            row4_val = match_row.iloc[0, 4]  # Column E
            ws_type.cell(row=3, column=col_idx, value=row3_val)
            ws_type.cell(row=4, column=col_idx, value=row4_val)
        else:
            ws_type.cell(row=3, column=col_idx, value="Not Found")
            ws_type.cell(row=4, column=col_idx, value="Not Found")

    # Save to BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output

# --- Streamlit UI ---
st.set_page_config(page_title="Template Automation", layout="wide")
st.title("📊 SKU Template Automation Tool")

st.markdown("Upload your **Input Excel file** below:")

input_file = st.file_uploader("Choose an input Excel file", type=["xlsx"])

if input_file:
    mapping_df = load_mapping()
    st.success("✅ Input file uploaded successfully.")
    
    if st.button("🔁 Generate Output File"):
        with st.spinner("Processing..."):
            result = process_file(input_file, mapping_df)
            st.success("✅ Output file generated successfully!")
            st.download_button(
                label="📥 Download Output Template",
                data=result,
                file_name="output_template.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )

st.markdown("---")
st.caption("Built for Rubick.ai | By Vishnu Sai")
